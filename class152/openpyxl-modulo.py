from random import uniform
import openpyxl
import os

BASE_DIR = os.path.dirname(os.path.realpath(__name__))

# carregando planilha
pedidos = openpyxl.load_workbook(f'{BASE_DIR}/class152/pedidos.xlsx')

qtd_paginas = pedidos.sheetnames # lista as paginas da planilha
print(qtd_paginas)

planilha1 = pedidos['Página1']
print(planilha1['b4'].value) # acessando cedula especifica

#----------------------------------------------------------------------------------------------------------------------------------------
# quando um valor é alterado, essa ação não é feita no arquivo original, é necessario salvar a
# planilha com a alteração em outro arquivo

planilha1['b4'].value = '12345' # alterando valor
pedidos.save(f'{BASE_DIR}/class152/nova-planilha.xlsx') # salvando

#----------------------------------------------------------------------------------------------------------------------------------------

for campo in planilha1['b']: # acessando coluna inteira 
    print(campo.value)


for linha in planilha1['a1:c2']: # acessando valores por range
    for coluna in linha:
        print(coluna.value)


for linha in planilha1: # exibindo planilha inteira
    print(linha[0]. value, linha[1]. value, linha[2]. value, linha[3]. value)

#----------------------------------------------------------------------------------------------------------------------------------------
# adicionando mais valores a planilha
# cell() - acessa colunas e linhas

for linha in range(5, 16):
    numero_pedido = linha -1
    planilha1.cell(linha, 1).value = '>'    
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

pedidos.save(f'{BASE_DIR}/class152/nova-planilha.xlsx') 

#----------------------------------------------------------------------------------------------------------------------------------------
# criando arquivo xlsx
# cell(linha, coluna) - acessa colunas e linhas por numerção (satrt=0)

m_planilha = openpyxl.Workbook()# criando arquivo xlsx

# Planilha2 - nome da planilha| 0 - indice onde estara localizada
m_planilha.create_sheet('Planilha2', 0)
m_planilha.create_sheet('Planilha3', 1)

planilha2 = m_planilha['Planilha2']
planilha3 = m_planilha['Planilha3']

for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha2.cell(linha, 1).value = numero_pedido
    planilha2.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha2.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilha2.cell(linha, 1).value = f'Gu {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 2).value = f'Fe {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(linha, 3).value = f'Li {linha} {round(uniform(10, 100), 2)}'



m_planilha.save(f'{BASE_DIR}/class152/minha-planilha.xlsx') 
